from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import uuid


def create_excel_from_table(table_data, output_path):
    """从表格数据创建Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "表格数据"
    
    # 获取表头和行数据
    headers = table_data.get('headers', [])
    rows = table_data.get('rows', [])
    
    # 样式定义（不使用背景颜色，保持简洁，所有文字黑色，不换行）
    # 明确设置黑色字体（000000是黑色）
    header_font = Font(bold=True, size=11, color="000000")  # 黑色加粗
    data_font = Font(size=11, color="000000")  # 黑色普通
    border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False, shrink_to_fit=False)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False, shrink_to_fit=False)
    
    # 写入表头
    if headers:
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment
            # 不设置 fill，使用默认的白色背景（无填充）
    
    # 写入数据行
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(cell_value) if cell_value is not None else "")
            cell.font = data_font
            cell.border = border
            cell.alignment = center_alignment  # 数据行也居中
            # 不设置 fill，使用默认的白色背景（无填充）
    
    # 自动调整列宽
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        # 检查表头长度
        if headers and col_idx <= len(headers):
            max_length = len(str(headers[col_idx - 1]))
        
        # 检查数据行长度
        for row in rows:
            if col_idx <= len(row):
                cell_value = str(row[col_idx - 1]) if row[col_idx - 1] is not None else ""
                max_length = max(max_length, len(cell_value))
        
        # 设置列宽（最小10，最大50）
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存文件
    wb.save(output_path)
    return output_path


def create_excel_from_multi_page(multi_page_data, output_path):
    """从多页表格数据创建Excel文件（每页一个工作表）"""
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认工作表
    
    pages = multi_page_data.get('pages', [])
    
    for page_idx, page_data in enumerate(pages):
        # 创建新工作表
        page_num = page_data.get('page', page_idx + 1)
        ws = wb.create_sheet(title=f"第{page_num}页")
        
        # 获取表头和行数据
        headers = page_data.get('headers', [])
        rows = page_data.get('rows', [])
        
        # 样式定义（不使用背景颜色，保持简洁，所有文字黑色，不换行）
        header_font = Font(bold=True, size=11, color="000000")  # 黑色加粗
        data_font = Font(size=11, color="000000")  # 黑色普通
        border = Border(
            left=Side(style='thin', color="000000"),
            right=Side(style='thin', color="000000"),
            top=Side(style='thin', color="000000"),
            bottom=Side(style='thin', color="000000")
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False, shrink_to_fit=False)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False, shrink_to_fit=False)
        
        # 写入表头
        if headers:
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=str(header))
                cell.font = header_font
                cell.border = border
                cell.alignment = center_alignment
                # 不设置 fill，使用默认的白色背景（无填充）
        
        # 写入数据行
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(cell_value) if cell_value is not None else "")
                cell.font = data_font
                cell.border = border
                cell.alignment = center_alignment  # 数据行也居中
                # 不设置 fill，使用默认的白色背景（无填充）
        
        # 自动调整列宽
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            if headers and col_idx <= len(headers):
                max_length = len(str(headers[col_idx - 1]))
            
            for row in rows:
                if col_idx <= len(row):
                    cell_value = str(row[col_idx - 1]) if row[col_idx - 1] is not None else ""
                    max_length = max(max_length, len(cell_value))
            
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存文件
    wb.save(output_path)
    return output_path

